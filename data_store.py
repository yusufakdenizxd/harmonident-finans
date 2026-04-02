import os
import sys
from pathlib import Path

if getattr(sys, 'frozen', False):
    if sys.platform == "darwin":
        base_path = Path(sys.executable).parent.parent
    else:
        base_path = Path(sys.executable).parent
    DATA_DIR = base_path / "data"
else:
    DATA_DIR = Path("data")

DATA_DIR.mkdir(parents=True, exist_ok=True)

from datetime import datetime
import json
import uuid
from dataclasses import dataclass, asdict
from typing import Optional, Any
from openpyxl import load_workbook
import xlrd

@dataclass
class Transaction:
    id: str
    datetime: str
    value: float
    description: str
    bank_name: str
    
    @property
    def is_income(self) -> bool:
        return self.value >= 0


@dataclass
class ImportError:
    row: int
    field: str
    message: str
    value: Any = None


class BankStore:
    def __init__(self, filename: str = "banks.json"):
        self.filepath = DATA_DIR / filename
    
    def load(self) -> list[str]:
        if not self.filepath.exists():
            return []
        with open(self.filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def save(self, banks: list[str]):
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump(banks, f, indent=2, ensure_ascii=False)
    
    def add(self, bank_name: str):
        banks = self.load()
        if bank_name not in banks:
            banks.append(bank_name)
            self.save(banks)
    
    def get_all(self) -> list[str]:
        return self.load()


@dataclass
class ImportLog:
    filename: str
    row_count: int
    datetime: str


class ImportLogStore:
    def __init__(self, filename: str = "import_log.json"):
        self.filepath = DATA_DIR / filename
    
    def load(self) -> list[ImportLog]:
        if not self.filepath.exists():
            return []
        with open(self.filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
            return [ImportLog(**item) for item in data]
    
    def save(self, logs: list[ImportLog]):
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump([asdict(log) for log in logs], f, indent=2, ensure_ascii=False)
    
    def add(self, log: ImportLog):
        logs = self.load()
        logs.append(log)
        self.save(logs)
    
    def get_all(self) -> list[ImportLog]:
        return self.load()


class ExcelImporter:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook = None
        self.worksheet = None
        self.columns: list[str] = []
        self.raw_columns: list[str] = []
        self.preview_data: list[dict] = []
        self.is_xls = filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx')
        self.col_indices: dict[str, int] = {}
        self.header_row_index: int = 0
    
    @property
    def default_bank_name(self) -> str:
        import os
        basename = os.path.basename(self.filepath)
        name_without_ext = os.path.splitext(basename)[0]
        return name_without_ext.strip()
    
    def load_file(self) -> bool:
        try:
            if self.is_xls:
                self.workbook = xlrd.open_workbook(self.filepath)
                self.worksheet = self.workbook.sheet_by_index(0)
                max_rows_to_check = min(5, self.worksheet.nrows)
                self.raw_columns, self.header_row_index = self._detect_header_row_xls(max_rows_to_check)
            else:
                self.workbook = load_workbook(filename=self.filepath, read_only=True, data_only=True)
                self.worksheet = self.workbook.active
                max_rows = self.worksheet.max_row if self.worksheet.max_row else 5
                max_rows = min(5, max_rows)
                rows_data = []
                for i in range(1, max_rows + 1):
                    rows_data.append([cell.value for cell in self.worksheet[i]])
                self.raw_columns, self.header_row_index = self._detect_header_row(rows_data)
            
            self.columns = [str(c) if c else "" for c in self.raw_columns]
            
            self.col_indices = {}
            filtered = []
            for idx, col in enumerate(self.columns):
                col = col.strip()
                if col:
                    self.col_indices[col] = idx
                    filtered.append(col)
            self.columns = filtered
            
            if self.is_xls:
                self._load_preview_xls()
            else:
                self._load_preview()
            
            return True
        except Exception as e:
            print(f"[ERROR] load_file failed: {e}")
            return False
    
    def _detect_header_row_xls(self, max_rows: int) -> tuple:
        best_row = []
        best_count = 0
        best_idx = 0
        
        for row_idx in range(max_rows):
            row_values = [self.worksheet.cell_value(row_idx, col) for col in range(self.worksheet.ncols)]
            non_empty = sum(1 for v in row_values if v is not None and str(v).strip())
            if non_empty > best_count:
                best_count = non_empty
                best_row = row_values
                best_idx = row_idx
        
        return best_row, best_idx
    
    def _detect_header_row(self, rows_data: list) -> tuple:
        best_row = []
        best_count = 0
        best_idx = 0
        
        for idx, row in enumerate(rows_data):
            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            if non_empty > best_count:
                best_count = non_empty
                best_row = row
                best_idx = idx
        
        return best_row, best_idx
    
    def _is_row_empty(self, row_dict: dict) -> bool:
        for val in row_dict.values():
            if val is not None and str(val).strip():
                return False
        return True
    
    def _load_preview_xls(self, max_rows: int = 10):
        self.preview_data = []
        for row_idx in range(1, min(self.worksheet.nrows, max_rows + 1)):
            row_dict = {}
            for col_name, col_idx in self.col_indices.items():
                row_dict[col_name] = self.worksheet.cell_value(row_idx, col_idx)
            self.preview_data.append(row_dict)
    
    def _load_preview(self, max_rows: int = 10):
        self.preview_data = []
        for row in self.worksheet.iter_rows(min_row=2, max_row=max_rows + 1, values_only=True):
            row_dict = {}
            for col_name, col_idx in self.col_indices.items():
                if col_idx < len(row):
                    row_dict[col_name] = row[col_idx]
            self.preview_data.append(row_dict)
    
    def validate_and_import(self, column_mapping: dict[str, str]) -> tuple[list[Transaction], list[ImportError]]:
        errors: list[ImportError] = []
        transactions: list[Transaction] = []
        
        start_row = self.header_row_index + 1
        
        if self.is_xls:
            for row_idx in range(start_row, self.worksheet.nrows):
                row_dict = {}
                for col_name, col_idx in self.col_indices.items():
                    row_dict[col_name] = self.worksheet.cell_value(row_idx, col_idx)
                
                if self._is_row_empty(row_dict):
                    continue
                
                transaction, row_errors = self._process_row(row_idx + 1, row_dict, column_mapping)
                if transaction:
                    transactions.append(transaction)
                errors.extend(row_errors)
                if row_errors:
                    for e in row_errors:
                        print(f"[ERROR] Row {e.row}, Field: {e.field}, Value: {e.value}, Message: {e.message}")
        else:
            for i, row in enumerate(self.worksheet.iter_rows(min_row=start_row + 1, values_only=True), start=start_row + 1):
                row_dict = {}
                for col_name, col_idx in self.col_indices.items():
                    if col_idx < len(row):
                        row_dict[col_name] = row[col_idx]
                
                if self._is_row_empty(row_dict):
                    continue
                
                transaction, row_errors = self._process_row(i, row_dict, column_mapping)
                if transaction:
                    transactions.append(transaction)
                errors.extend(row_errors)
                if row_errors:
                    for e in row_errors:
                        print(f"[ERROR] Row {e.row}, Field: {e.field}, Value: {e.value}, Message: {e.message}")
        
        return transactions, errors
    
    def _process_row(self, row_num: int, row_data: dict, mapping: dict[str, str]) -> tuple[Optional[Transaction], list[ImportError]]:
        errors: list[ImportError] = []
        
        datetime_val = row_data.get(mapping.get("datetime"))
        if datetime_val is None:
            errors.append(ImportError(row=row_num, field="datetime", message="Tarih boş", value=datetime_val))
            return None, errors
        
        try:
            if isinstance(datetime_val, datetime):
                dt = datetime_val
            elif isinstance(datetime_val, (int, float)):
                dt = xlrd.xldate.xldate_as_datetime(datetime_val, self.workbook.datemode)
            elif isinstance(datetime_val, str):
                date_str = datetime_val.strip()
                date_str = date_str.replace(".", "-").replace("/", "-")
                try:
                    dt = datetime.strptime(date_str, "%d-%m-%Y %H:%M")
                except ValueError:
                    dt = datetime.strptime(date_str, "%d-%m-%Y")
            else:
                date_str = str(datetime_val).strip()
                date_str = date_str.replace(".", "-").replace("/", "-")
                try:
                    dt = datetime.strptime(date_str, "%d-%m-%Y %H:%M")
                except ValueError:
                    dt = datetime.strptime(date_str, "%d-%m-%Y")
            formatted_datetime = dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            errors.append(ImportError(row=row_num, field="datetime", message=f"Geçersiz tarih (DD-MM-YYYY olmalı): {e}", value=datetime_val))
            return None, errors
        
        value_raw = row_data.get(mapping.get("value"))
        try:
            if value_raw is None or str(value_raw).strip() == "":
                errors.append(ImportError(row=row_num, field="value", message="Tutar boş", value=value_raw))
                return None, errors
            value = float(str(value_raw).replace(",", "."))
        except ValueError:
            errors.append(ImportError(row=row_num, field="value", message="Geçersiz sayı", value=value_raw))
            return None, errors
        
        description = str(row_data.get(mapping.get("description"), "")).strip()
        
        bank_name = str(mapping.get("bank_name") or "")
        
        if not errors:
            transaction = Transaction(
                id=str(uuid.uuid4()),
                datetime=formatted_datetime,
                value=value,
                description=description,
                bank_name=bank_name
            )
            return transaction, []
        
        return None, errors


class DataStore:
    def __init__(self, filename: str = "transactions.json"):
        self.filepath = DATA_DIR / filename
    
    def load(self) -> list[Transaction]:
        if not self.filepath.exists():
            return []
        
        with open(self.filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        return [Transaction(**t) for t in data]
    
    def save(self, transactions: list[Transaction]):
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump([asdict(t) for t in transactions], f, indent=2, ensure_ascii=False)
    
    def add(self, transaction: Transaction):
        transactions = self.load()
        transactions.append(transaction)
        self.save(transactions)
    
    def get_by_date_range(self, start: str, end: str, bank: str = None) -> list[Transaction]:
        transactions = self.load()
        filtered = [t for t in transactions if start <= t.datetime[:10] <= end]
        if bank and bank != "Tümü":
            filtered = [t for t in filtered if t.bank_name == bank]
        return filtered
    
    def get_summary(self, start: str, end: str, bank: str = None) -> dict:
        transactions = self.get_by_date_range(start, end, bank)
        
        total_income = sum(t.value for t in transactions if t.value > 0)
        total_expense = sum(t.value for t in transactions if t.value < 0)
        net_balance = total_income + total_expense
        
        return {
            "total_income": total_income,
            "total_expense": abs(total_expense),
            "net_balance": net_balance,
            "count": len(transactions)
        }
